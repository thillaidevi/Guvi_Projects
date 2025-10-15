import openpyxl
import os

# Absolute path to credentials Excel file
EXCEL_PATH = os.path.join(
    "C:\\Users\\Rajesh C\\PycharmProjects\\Guvi_Projects\\project_orangehrm_playwright\\data",
    "credentials.xlsx"
)

def get_login_data():
    """
    Reads login credentials and expected test outcome from Excel.
    Returns a list of tuples: (si_no, username, password, expected)
    Columns:
        A = Serial Number
        F = Username
        G = Password
        H = Expected Result ("pass"/"fail")
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active
        data = []

        # Skip header row and extract values only
        for row in sheet.iter_rows(min_row=2, values_only=True):
            si_no = row[0]              # Column A

            username = row[5]           # Column F

            password = row[6]           # Column G

            expected = row[7].strip().lower() if row[7] else ""  # Column H

            data.append((si_no, username, password, expected))
        wb.close()
        return data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def write_result(si_no, result):
    """
        Writes the test result back to Excel for the given serial number.
        Updates Column I (index 8) with the result string.
    """

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active

        # Locate row by serial number and update result column
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == si_no:
                row[8].value = result  # Column I = Test Result
                break
        wb.save(EXCEL_PATH)
        wb.close()
    except Exception as e:
        print(f"Error writing to Excel file: {e}")